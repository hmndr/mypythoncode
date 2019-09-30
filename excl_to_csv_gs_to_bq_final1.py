import xlrd
import openpyxl
import os
import re
import sys
import csv
import subprocess
reload(sys)
sys.setdefaultencoding('utf-8')



argument = sys.argv


bucket_path = 'gs://bucketh10/'

Archive_path = 'gs://bucketh10/Archive1/'

VM_path = '//home/user1/prod_script/'

#Dataset = 'test_dataset'

Table_apac = 'mattel_vendor_report_APAC'

Table_emea = 'mattel_vendor_report_EMEA'

Table_nad = 'mattel_vendor_report_NAD'



cmd = 'gsutil ls -l ' + bucket_path + '> out.txt'
os.system(cmd)

file_name = argument[1]

Project_name = argument[2]

Dataset = argument[3]

file_object = open('out.txt')

Flag=0

for line in file_object:
    match_obj = re.match( '.*  %s(.*)' % (bucket_path), line)
    if(match_obj):
        File = match_obj.group(1)
        if file_name.lower() in File.lower():
            Raw_File_Name = File
            Flag=1

if Flag==0:
    exit()


Excel_File_Name=Raw_File_Name.replace(" ","_")

cmd1 = 'gsutil cp '+bucket_path+'''"'''+Raw_File_Name+'''"'''+' '+VM_path+Excel_File_Name
os.system(cmd1)




if "asia" in Excel_File_Name.lower():

    CSV_File = "hwid_mattel_vender_report_apac.csv"
    
    cmd2 = 'date +"%m-%d-%Y"'
    Date2= subprocess.check_output(cmd2,shell=True).strip()

    wb = xlrd.open_workbook(VM_path+Excel_File_Name)
    sh1 = wb.sheet_by_index(0)
    col = sh1.ncols
    row = sh1.nrows
    Excel_File = openpyxl.load_workbook(VM_path+Excel_File_Name)

    sheet = Excel_File.active
    sheet.cell(row=1,column=col+1).value = 'LOAD_DTM'
    for i in range(2,row+1):
        sheet.cell(row=i,column=col+1).value = Date2

    Excel_File.save(Excel_File_Name)


    wb = xlrd.open_workbook(VM_path+Excel_File_Name)
    sh = wb.sheet_by_index(0)
    your_csv_file = open(VM_path+"hwid_mattel_vender_report_apac.csv", 'w')
    wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

    for rownum in range(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    your_csv_file.close()

    Date_File = "hwid_apac_date.txt"

    cmd3='touch hwid_apac_date.txt'
    os.system(cmd3)

    date_writer = open(VM_path+"hwid_apac_date.txt",'w')
    Date_list=Excel_File_Name.split('_')
    Date = Date_list[-1].split('.')
    date_writer.write("date"+"\n")
    Date_Rev = Date[0]
    date_writer.write(Date_Rev[4]+Date_Rev[5]+Date_Rev[6]+Date_Rev[7]+Date_Rev[2]+Date_Rev[3]+Date_Rev[0]+Date_Rev[1]+"\n")
    date_writer.close()
    
    
    cmd4 = 'gsutil mv '+VM_path+CSV_File+' '+bucket_path
    os.system(cmd4)

    cmd5 = 'bq load --autodetect --source_format=CSV --project_id='+Project_name+' '+Dataset+'.'+Table_apac+' '+bucket_path+"hwid_mattel_vender_report_apac.csv" 
    os.system(cmd5)	
    
    os.system('gsutil mv '+bucket_path+'''"'''+Raw_File_Name+'''"'''+' '+Archive_path+'''"'''+Raw_File_Name+'''"''')
   



elif "emea" in Excel_File_Name.lower():

    CSV_File = 'hwid_mattel_vender_report_emea.csv'


    cmd7 = 'date +"%m-%d-%Y"'
    Date3= subprocess.check_output(cmd8,shell=True).strip()


    cmd1 = 'date +"%m-%d-%Y"'
    Date2= subprocess.check_output(cmd1,shell=True).strip()

    wb = xlrd.open_workbook(Excel_File_Name)
    sh1 = wb.sheet_by_index(0)
    col = sh1.ncols
    row = sh1.nrows
    Excel_File = openpyxl.load_workbook(Excel_File_Name)

    sheet = Excel_File.active
    sheet.cell(row=2,column=col+1).value = 'LOAD_DTM'
    for i in range(3,row+1):
        sheet.cell(row=i,column=col+1).value = Date2

    Excel_File.save(Excel_File_Name)



    xfile = openpyxl.load_workbook(VM_path+Excel_File_Name)
    sheet = xfile.active
    Date_list= sheet.cell(row = 2, column = 52)
    date_list=Date_list.value

    sheet['AZ2'] = 'All_STi_Week_1'
    sheet['AY2'] = 'All_ST_Week_2'
    sheet['AX2'] = 'All_ST_Week_3'
    sheet['AW2'] = 'All_ST_Week_4'
    sheet['J2'] = 'Bookings_Week4'
    sheet['K2'] = 'Bookings_Week3'
    sheet['L2'] = 'Bookings_Week2'
    sheet['M2'] = 'Bookings_Week1'
    sheet['Q2'] = 'Fcst_Week1'
    sheet['R2'] = 'Fcst_Week2'
    sheet['S2'] = 'Fcst_Week3'
    sheet['T2'] = 'Fcst_Week4'
    sheet['U2'] = 'Fcst_Week5'
    sheet['V2'] = 'Fcst_Week6' 
    sheet['W2'] = 'Fcst_Week7'
    sheet['X2'] = 'Fcst_Week8'
    sheet['Y2'] = 'Fcst_Week9'
    sheet['Z2'] = 'Fcst_Week10'
    sheet['AA2'] = 'Fcst_Week11'
    sheet['AB2'] = 'Fcst_Week12'
    sheet['AC2'] = 'Fcst_Week13'
    sheet['AE2'] = 'Retail_ST_Week4'
    sheet['AF2'] = 'Retail_ST_Week3'
    sheet['AG2'] = 'Retail_ST_Week2'
    sheet['AH2'] = 'Retail_ST_Week1'
    sheet['AK2'] = 'Online_ST_Week4'
    sheet['AL2'] = 'Online_ST_Week3'
    sheet['AM2'] = 'Online_ST_Week2'
    sheet['AN2'] = 'Online_ST_Week1'	
    sheet['AQ2'] = 'Rslr_ST_Week4'
    sheet['AR2'] = 'Rslr_ST_Week3'
    sheet['AS2'] = 'Rslr_ST_Week2'
    sheet['AT2'] = 'Rslr_ST_Week1'
    
    xfile.save(VM_path+Excel_File_Name)

    wb = xlrd.open_workbook(VM_path+Excel_File_Name)
    sh = wb.sheet_by_index(0)
    your_csv_file = open(VM_path+'hwid_mattel_vender_report_emea.csv', 'w')
    wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

    for rownum in range(1,sh.nrows):
        wr.writerow(sh.row_values(rownum))

    your_csv_file.close()

    Date_separated= date_list.split("_")

    Date = Date_separated[2].split("-")

    Date_File = "hwid_emea_date.txt"

    cmd2='touch hwid_emea_date.txt'
    os.system(cmd2)

    date_writer = open(VM_path+"hwid_emea_date.txt",'w')
    date_writer.write("date"+"\n")
    date_writer.write(Date[0]+Date[1]+Date[2]+"\n")
    date_writer.close()
   
    File_Name = Excel_File_Name.split(".")

    cmd3='gsutil mv '+bucket_path+Excel_File_Name+' '+bucket_path+File_Name[0]+'_'+Date3+'.'+File_Name[1]
    os.system(cmd3)

    cmd4='mv '+VM_path+Excel_File_Name+' '+VM_path+File_Name[0]+'_'+Date3+'.'+File_Name[1]
    os.system(cmd4)


    
    Excel_File_Name=File_Name[0]+'_'+Date3+'.'+File_Name[1]

    cmd5 = 'gsutil mv '+VM_path+CSV_File+' '+bucket_path
    os.system(cmd5)

    cmd6 = 'bq load --autodetect --source_format=CSV --project_id='+Project_name+' '+Dataset+'.'+Table_emea+' '+bucket_path+'hwid_mattel_vender_report_emea.csv'
    os.system(cmd6)
    
    os.system('gsutil mv '+bucket_path+Excel_File_Name+' '+Archive_path+Excel_File_Name)
    
    

else:

    CSV_File = 'hwid_mattel_vender_report_nad.csv'

    cmd1 = 'date +"%m-%d-%Y"'
    Date2= subprocess.check_output(cmd1,shell=True).strip()

    wb = xlrd.open_workbook(Excel_File_Name)
    sh1 = wb.sheet_by_index(0)
    col = sh1.ncols
    row = sh1.nrows
    Excel_File = openpyxl.load_workbook(Excel_File_Name)

    sheet = Excel_File.active
    sheet.cell(row=1,column=col+1).value = 'LOAD_DTM'
    for i in range(2,row+1):
        sheet.cell(row=i,column=col+1).value = Date2

    Excel_File.save(Excel_File_Name)

    wb = xlrd.open_workbook(VM_path+Excel_File_Name)
    sh = wb.sheet_by_index(0)
    your_csv_file = open(VM_path+'hwid_mattel_vender_report_nad.csv', 'w')
    wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

    for rownum in range(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    your_csv_file.close()

    Date_File = "hwid_nad_date.txt"

    cmd2='touch hwid_nad_date.txt'
    os.system(cmd2)

    date_writer = open(VM_path+"hwid_nad_date.txt",'w')
    date_list=Excel_File_Name.split('.')
    Date=date_list[0].split('_')
    date_writer.write("date"+"\n")
    date_writer.write(Date[1]+Date[2]+Date[3]+"\n")
    date_writer.close()
    
    
    cmd3 = 'gsutil mv '+VM_path+CSV_File+' '+bucket_path
    os.system(cmd3)

    cmd4 = 'bq load --autodetect --source_format=CSV --project_id='+Project_name+' '+Dataset+'.'+Table_nad+' '+bucket_path+'hwid_mattel_vender_report_nad.csv' 
    os.system(cmd4)	

    os.system('gsutil mv '+bucket_path+Excel_File_Name+' '+Archive_path+Excel_File_Name)


#else:

 #   print ("Provided Input File Not Present On Location")


os.system('gsutil mv '+VM_path+Date_File+' ' +bucket_path)

#os.system('gsutil mv '+bucket_path+'''"'''+Raw_File_Name+'''"'''+' '+Archive_path+'''"'''+Raw_File_Name+'''"''')

#os.system('gsutil mv '+bucket_path+Excel_File_Name+' '+Archive_path+Excel_File_Name)

os.system('rm '+VM_path+Excel_File_Name)

os.system('rm out.txt')


